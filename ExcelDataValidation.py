import openpyxl

# Função para obter o índice da coluna a partir da letra da coluna (A=1, B=2, etc.)
def obter_indice_coluna_da_letra(letra):
    letra = letra.upper()
    alfabeto = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    indice_coluna = 0

    for i in range(len(letra)):
        caracter = letra[i]
        indice_caracter = alfabeto.index(caracter) + 1
        indice_coluna = indice_coluna * 26 + indice_caracter

    return indice_coluna

# Caminho para o arquivo Excel
caminho_arquivo_excel = r'C:\Users\jalmeida26\OneDrive - DXC Production\Desktop\Projeto ValidacaoDados\WorkFile.xlsx'

# Carregar o arquivo Excel
try:
    workbook = openpyxl.load_workbook(caminho_arquivo_excel)
    sheet = workbook['Devices']  # Nome da folha
except Exception as e:
    print(f'Erro ao abrir o arquivo Excel: {e}')
    exit()

# Especificar as colunas a verificar (com base nas letras das colunas)
colunas_a_verificar = ["B", "C", "D", "H", "I", "L", "M", "N", "O", "Q", "R", "T", "U", "Y", "Z", "AA", "AD", "AH", "AI", "AJ", "AK", "AM", "AN", "AO"]

# Especificar as linhas a verificar
linhas_a_verificar = range(6, 11)  # Exemplo: Verificar linhas de 6 a 11 = 6, 11

# Especificar a linha com os nomes
linha_com_nomes = 1  # Atualizar com o número da linha que contém os nomes das colunas

# Inicializar uma lista para armazenar as mensagens das colunas vazias
mensagens_colunas_vazias = []

# Percorrer colunas e linhas e obter o nome da coluna na linha especificada
for letra_coluna in colunas_a_verificar:
    indice_coluna = obter_indice_coluna_da_letra(letra_coluna)

    if indice_coluna == 0:
        print(f'Letra de coluna inválida: {letra_coluna}')
        continue

    # Obter o nome da coluna
    nome_coluna = sheet.cell(row=linha_com_nomes, column=indice_coluna).value

    # Percorrer linhas e verificar dados em falta
    for linha in linhas_a_verificar:
        valor_celula = sheet.cell(row=linha, column=indice_coluna).value

        if valor_celula is None:
            mensagem = f'Linha: {linha} - Dados em falta na Coluna: {nome_coluna}'
            print(mensagem)
            mensagens_colunas_vazias.append(mensagem)

# Criar o arquivo .txt com as mensagens das colunas vazias
caminho_arquivo_txt = r'C:\Users\jalmeida26\OneDrive - DXC Production\Desktop\DadosEmFaltaExcel.txt'
with open(caminho_arquivo_txt, 'w') as arquivo_txt:
    for mensagem in mensagens_colunas_vazias:
        arquivo_txt.write(mensagem + '\n')

# Fechar o arquivo Excel
workbook.close()
